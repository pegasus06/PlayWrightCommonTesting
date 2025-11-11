import os.path

import xlrd
import openpyxl


class ExcelUtil:
    def __init__(self, excel_path):
        self.data = xlrd.open_workbook(excel_path)
        self.table = None
        self.keys = None
        self.row_num = None
        self.col_num = None

    def set_sheet_name(self, sheet_name):
        self.table = self.data.sheet_by_name(sheet_name)
        self.keys = self.table.row_values(0)
        self.row_num = self.table.nrows
        self.col_num = self.table.ncols

    def get_datas(self, case_name):
        """根据用例名case name获取对应的测试数据
        :param case_name:当前测试用例方法名
        :return: 字典格式的测试数据的集合
        """

        datas = []
        for i in range(1, self.row_num):
            if case_name in self.table.cell_value(i, 0):
                row_data = {}
                for j in range(0, self.col_num):
                    row_data[self.keys[j]] = self.table.cell_value(i, j)
                datas.append(row_data)
        return datas

    def get_datas_by_subtitle(self, case_name, subtitle):
        """
        根据当前测试用例方法名case name和过滤条件filter筛选Excel中的数据
        :param case_name: 用例名
        :param subtitle: excel中“subtitle”字段的值(当同一个用例存在多组数据时，通过副标题区分)
        :return:
        """
        subtitle_index = self.keys.index('subtitle')
        datas = {}
        for i in range(1, self.row_num):
            if case_name in self.table.cell_value(i, 0) and subtitle == self.table.cell_value(i, subtitle_index):
                for j in range(0, self.col_num):
                    datas[self.keys[j]] = self.table.cell_value(i, j)
                break
        return datas

    @staticmethod
    def new_or_append_to_excel(excel_path, sheet_name: str, data: dict):
        """
        创建或者追加数据给指定的excel对应的sheet页
        只支持后缀为.xLsx的新版本excel
        :param excel_path: excel文件的绝对路径 表名称xlsx
        :param sheet_name: 表名称
        :param data:表头:值 组成的字典
        """
        # 判断文件是否存在并创建部分:
        if sheet_name.lower() == "sheet":
            sheet_name = "sheet"
        if not os.path.exists(excel_path):
            wb = openpyxl.Workbook()
            if sheet_name == "Sheet":
                sheet = wb.active
            else:
                sheet = wb.create_sheet(sheet_name)
        else:
            wb = openpyxl.load_workbook(excel_path)
            if sheet_name not in wb.sheetnames:
                sheet = wb.create_sheet(sheet_name)
            else:
                wb = openpyxl.load_workbook(excel_path)
                sheet = wb[sheet_name]
        # 处理数据部分:
        # 获取表头内容:
        head = []
        # 获取要填写的行
        value_row = sheet.max_row + 1
        for row in sheet[1]:
            head.append(row.value)
        for key, value in data.items():
            if key not in head:
                # 此处由于max_column至少为1,所以等于空出了第一列
                sheet.cell(row=1, column=sheet.max_column + 1).value = key
                sheet.cell(row=2, column=sheet.max_column).value = value
            else:
                # 这里+1是因为下标从0开始
                key_column = head.index(key) + 1
                sheet.cell(row=value_row, column=key_column).value = value
        wb.save(excel_path)
